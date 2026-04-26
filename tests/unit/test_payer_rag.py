from __future__ import annotations

import json
import tempfile
import unittest
from pathlib import Path
import shutil
import uuid

from openpyxl import load_workbook

from lam.payer_rag.analyze import analyze_outliers
from lam.payer_rag.export import build_validation_queue_rows, export_workbook, write_csv
from lam.payer_rag.ingest import default_source_manifest, parse_duke_standard_charges_text, parse_shoppable_services_filters_text
from lam.payer_rag.rag import ask_question, build_index
from lam.payer_rag.sample_data import SAMPLE_DUKE_STANDARD_CHARGES
from lam.payer_rag.workflow import CurrentTaskContract, build_workspace, ensure_workspace, extract_current_task_contract


class PayerRagTests(unittest.TestCase):
    def make_workspace(self) -> Path:
        root = Path("test_artifacts") / f"payer_rag_{uuid.uuid4().hex}"
        (root / "normalized").mkdir(parents=True, exist_ok=True)
        (root / "artifacts").mkdir(parents=True, exist_ok=True)
        self.addCleanup(lambda: shutil.rmtree(root, ignore_errors=True))
        return root

    def test_parse_duke_standard_charges_preserves_sources_and_handles_rates(self) -> None:
        bundle = parse_duke_standard_charges_text(
            SAMPLE_DUKE_STANDARD_CHARGES,
            source_name="sample source",
            source_url="sample://duke",
            geography="Durham, NC",
            service_keywords=("mri", "emergency"),
            max_services=10,
        )
        self.assertGreaterEqual(len(bundle["services"]), 2)
        self.assertTrue(all(row["source_url"] == "sample://duke" for row in bundle["services"]))
        rate = bundle["rates"][0]
        self.assertIsInstance(rate["negotiated_rate"], float)
        self.assertEqual(rate["source_url"], "sample://duke")
        self.assertIn("row_", rate["raw_reference"])

    def test_parse_duke_standard_charges_balances_keyword_coverage(self) -> None:
        bundle = parse_duke_standard_charges_text(
            SAMPLE_DUKE_STANDARD_CHARGES,
            source_name="sample source",
            source_url="sample://duke",
            geography="Durham, NC",
            service_keywords=("mri", "acute major eye"),
            max_services=2,
        )
        descriptions = {row["description"].lower() for row in bundle["services"]}
        self.assertIn("mri brain without contrast", descriptions)
        self.assertIn("acute major eye infections without cc/mcc", descriptions)

    def test_analyze_outliers_flags_high_variance(self) -> None:
        bundle = parse_duke_standard_charges_text(
            SAMPLE_DUKE_STANDARD_CHARGES,
            source_name="sample source",
            source_url="sample://duke",
            geography="Durham, NC",
            service_keywords=("mri", "colonoscopy", "emergency", "acute major eye"),
            max_services=10,
        )
        results = analyze_outliers(
            rates=bundle["rates"],
            plans=bundle["plans"],
            payers=bundle["payers"],
            services=bundle["services"],
            outlier_threshold=0.15,
            min_peer_count=3,
        )
        self.assertTrue(results["candidates"])
        self.assertTrue(any("peer median" in row["reason_flagged"] for row in results["candidates"]))

    def test_analyze_outliers_excludes_supply_like_rc_rows(self) -> None:
        results = analyze_outliers(
            rates=[
                {
                    "rate_id": "1",
                    "payer_id": "payer_a",
                    "plan_id": "plan_a",
                    "service_id": "service_supply",
                    "facility_name": "Duke",
                    "setting": "outpatient",
                    "billing_class": "facility",
                    "negotiated_rate": 500.0,
                    "source_url": "sample://duke",
                    "raw_reference": "row_1",
                    "confidence": 0.9,
                },
                {
                    "rate_id": "2",
                    "payer_id": "payer_b",
                    "plan_id": "plan_b",
                    "service_id": "service_supply",
                    "facility_name": "Duke",
                    "setting": "outpatient",
                    "billing_class": "facility",
                    "negotiated_rate": 100.0,
                    "source_url": "sample://duke",
                    "raw_reference": "row_1",
                    "confidence": 0.9,
                },
                {
                    "rate_id": "3",
                    "payer_id": "payer_c",
                    "plan_id": "plan_c",
                    "service_id": "service_supply",
                    "facility_name": "Duke",
                    "setting": "outpatient",
                    "billing_class": "facility",
                    "negotiated_rate": 90.0,
                    "source_url": "sample://duke",
                    "raw_reference": "row_1",
                    "confidence": 0.9,
                },
            ],
            plans=[
                {"plan_id": "plan_a", "payer_id": "payer_a", "plan_name": "A", "source_url": "sample://duke"},
                {"plan_id": "plan_b", "payer_id": "payer_b", "plan_name": "B", "source_url": "sample://duke"},
                {"plan_id": "plan_c", "payer_id": "payer_c", "plan_name": "C", "source_url": "sample://duke"},
            ],
            payers=[
                {"payer_id": "payer_a", "payer_name": "Aetna"},
                {"payer_id": "payer_b", "payer_name": "BCBS"},
                {"payer_id": "payer_c", "payer_name": "Cigna"},
            ],
            services=[
                {
                    "service_id": "service_supply",
                    "code_type": "RC",
                    "code": "0272",
                    "description": "Cover X-Ray C-Armor Tube (Ea)",
                }
            ],
            outlier_threshold=0.2,
            min_peer_count=3,
        )
        self.assertFalse(results["candidates"])

    def test_default_manifest_includes_regional_sources_and_tic_references(self) -> None:
        manifest = default_source_manifest()
        names = {row["source_name"] for row in manifest}
        self.assertIn("WakeMed Raleigh Campus and North Hospital standard charges", names)
        self.assertIn("UNC Health standard charges landing page", names)
        self.assertIn("Blue Cross NC transparency in coverage page", names)

    def test_parse_shoppable_services_filters_extracts_consumer_friendly_services(self) -> None:
        payload = """
        {
          "serviceLine": [
            {"label": "MRI Brain without contrast", "value": "70551"},
            {"label": "CT Abdomen and Pelvis with contrast", "value": "74177"}
          ]
        }
        """
        bundle = parse_shoppable_services_filters_text(
            payload,
            source_name="duke shoppable",
            source_url="https://example.com/shoppable",
            geography="Durham, NC",
        )
        descriptions = {row["description"] for row in bundle["services"]}
        self.assertIn("MRI Brain without contrast", descriptions)
        self.assertIn("CT Abdomen and Pelvis with contrast", descriptions)

    def test_build_validation_queue_rows_marks_claim_review_as_pending(self) -> None:
        rows = build_validation_queue_rows(
            [
                {
                    "priority_rank": 1,
                    "payer_name": "United Healthcare",
                    "plan_name": "Commercial/EPO/PPO",
                    "service": "MRI brain without contrast",
                    "code": "70551",
                    "payer_rate": 1725.0,
                    "peer_median": 1510.0,
                    "variance_percent": 0.142,
                    "source_evidence": "sample://duke | row_4",
                }
            ]
        )
        self.assertEqual(rows[0]["validation_status"], "pending_contract_review")
        self.assertEqual(rows[0]["sample_claim_needed"], "yes")

    def test_extract_current_task_contract_for_fairfax(self) -> None:
        contract = extract_current_task_contract(
            "Build a payer pricing package for Fairfax, VA with a RAG index, spreadsheet, source manifest, and stakeholder summary."
        )
        self.assertEqual(contract.geography, "Fairfax, VA")
        self.assertEqual(contract.market, "Fairfax")
        self.assertEqual(contract.state, "VA")
        self.assertTrue(contract.geography_explicit)
        self.assertIn("payer outreach spreadsheet", contract.requested_outputs)
        self.assertIn("fairfax_va_payer_outreach_candidates.xlsx", contract.required_artifacts)

    def test_build_workspace_writes_contract_and_geography_named_artifacts(self) -> None:
        root = self.make_workspace()
        contract = CurrentTaskContract(
            geography="Durham, NC",
            state="NC",
            market="Durham",
            requested_outputs=["stakeholder summary", "payer outreach spreadsheet", "RAG/vector store"],
            source_constraints=["public only"],
            required_artifacts=[],
        )
        result = build_workspace(contract=contract, workspace=root, offline_fallback=True)
        artifacts = result["artifact_paths"]
        self.assertTrue(Path(artifacts["workbook_xlsx"]).exists())
        self.assertIn("durham_nc_payer_outreach_candidates.xlsx", artifacts["workbook_xlsx"])
        self.assertTrue(Path(artifacts["task_contract_json"]).exists())
        validation = result.get("geography_validation", {})
        self.assertTrue(validation.get("passed"))

    def test_ensure_workspace_invalidates_different_geography_runs(self) -> None:
        stale_root = Path("data/payer_rag_runs") / f"durham_nc_{uuid.uuid4().hex}"
        stale_root.mkdir(parents=True, exist_ok=True)
        self.addCleanup(lambda: shutil.rmtree(stale_root, ignore_errors=True))
        (stale_root / "artifacts" / "rag_index").mkdir(parents=True, exist_ok=True)
        (stale_root / "normalized").mkdir(parents=True, exist_ok=True)
        (stale_root / "artifacts" / "durham_nc_payer_dashboard.html").write_text("<html>Durham, NC</html>", encoding="utf-8")
        (stale_root / "artifacts" / "durham_nc_payer_outreach_candidates.xlsx").write_text("", encoding="utf-8")
        (stale_root / "artifacts" / "durham_nc_summary_report.md").write_text("Durham, NC", encoding="utf-8")
        (stale_root / "artifacts" / "durham_nc_source_manifest.csv").write_text("source_name,source_type,source_url_or_path,accessed_or_ingested_date,geography,notes,confidence\nsample,synthetic_fixture,sample://duke,2026-04-25,\"Durham, NC\",fixture,0.4\n", encoding="utf-8")
        (stale_root / "artifacts" / "durham_nc_data_quality_report.md").write_text("Durham, NC", encoding="utf-8")
        (stale_root / "artifacts" / "durham_nc_outreach_candidates.csv").write_text("priority_rank\n", encoding="utf-8")
        (stale_root / "artifacts" / "durham_nc_payer_summary.csv").write_text("payer_name\n", encoding="utf-8")
        (stale_root / "artifacts" / "durham_nc_contract_validation_queue.csv").write_text("priority_rank\n", encoding="utf-8")
        (stale_root / "artifacts" / "durham_nc_validation_checklist.md").write_text("Durham, NC", encoding="utf-8")
        (stale_root / "artifacts" / "rag_index" / "payer_rag.db").write_text("", encoding="utf-8")
        (stale_root / "normalized" / "payers.csv").write_text("payer_id,payer_name\n", encoding="utf-8")
        (stale_root / "normalized" / "plans.csv").write_text("plan_id,payer_id\n", encoding="utf-8")
        (stale_root / "normalized" / "services.csv").write_text("service_id,description\n", encoding="utf-8")
        (stale_root / "normalized" / "rates.csv").write_text("rate_id,payer_id\n", encoding="utf-8")
        (stale_root / "task_contract.json").write_text(
            json.dumps(
                {
                    "geography": "Durham, NC",
                    "state": "NC",
                    "market": "Durham",
                    "domain": "insurance payer/plan pricing analysis",
                    "requested_outputs": ["stakeholder summary"],
                    "source_constraints": ["public only"],
                    "timeframe": "current public data",
                    "required_artifacts": [],
                    "stakeholder_audience": "payer contracting and healthcare analytics stakeholders",
                }
            ),
            encoding="utf-8",
        )
        contract = extract_current_task_contract("Build a payer pricing package for Fairfax, VA.")
        result = ensure_workspace(contract=contract, allow_reuse=True, offline_fallback=True)
        self.assertFalse(result.get("reused_existing_outputs"))
        self.assertTrue(any("Durham, NC" in item for item in result.get("invalidated_artifacts", [])))

    def test_rag_answers_include_sources(self) -> None:
        bundle = parse_duke_standard_charges_text(
            SAMPLE_DUKE_STANDARD_CHARGES,
            source_name="sample source",
            source_url="sample://duke",
            geography="Durham, NC",
            service_keywords=("mri", "colonoscopy", "emergency", "acute major eye"),
            max_services=10,
        )
        results = analyze_outliers(
            rates=bundle["rates"],
            plans=bundle["plans"],
            payers=bundle["payers"],
            services=bundle["services"],
            outlier_threshold=0.15,
            min_peer_count=3,
        )
        root = self.make_workspace()
        write_csv(root / "normalized" / "payers.csv", bundle["payers"], ["payer_id", "payer_name", "source_url", "source_type", "confidence", "notes"])
        write_csv(
            root / "normalized" / "plans.csv",
            bundle["plans"],
            ["plan_id", "payer_id", "plan_name", "plan_type", "market_segment", "geography", "network_name", "source_url", "confidence", "notes"],
        )
        write_csv(
            root / "normalized" / "services.csv",
            bundle["services"],
            ["service_id", "code_type", "code", "description", "category", "source_url"],
        )
        write_csv(
            root / "normalized" / "rates.csv",
            bundle["rates"],
            [
                "rate_id",
                "payer_id",
                "plan_id",
                "service_id",
                "provider_name",
                "provider_npi",
                "facility_name",
                "negotiated_rate",
                "allowed_amount",
                "cash_price",
                "billing_class",
                "setting",
                "geography",
                "effective_date",
                "source_url",
                "raw_reference",
                "confidence",
                "methodology",
                "percentile_count",
                "gross_charge",
            ],
        )
        write_csv(
            root / "artifacts" / "source_manifest.csv",
            [
                {
                    "source_name": "sample source",
                    "source_type": "synthetic_fixture",
                    "source_url_or_path": "sample://duke",
                    "accessed_or_ingested_date": "2026-04-25",
                    "notes": "fixture",
                }
            ],
            ["source_name", "source_type", "source_url_or_path", "accessed_or_ingested_date", "notes"],
        )
        write_csv(
            root / "artifacts" / "outreach_candidates.csv",
            results["candidates"],
            [
                "priority_rank",
                "payer_name",
                "plan_name",
                "service",
                "code",
                "payer_rate",
                "peer_median",
                "peer_min",
                "peer_max",
                "variance_percent",
                "reason_flagged",
                "confidence",
                "recommended_action",
                "source_evidence",
                "compared_service",
                "facility_name",
            ],
        )
        build_index(root)
        response = ask_question(root, "Which plans need outreach?")
        self.assertIn("sources", response)
        self.assertTrue(response["sources"])
        payer_response = ask_question(root, "Show evidence for why United Healthcare was flagged.")
        self.assertIn("United Healthcare", payer_response["answer"])
        no_match_response = ask_question(root, "Which plans are most expensive for ultrasound in Durham?")
        self.assertIn("No exact service match", no_match_response["answer"])
        synonym_response = ask_question(root, "Which plans are most expensive for magnetic resonance imaging in Durham?")
        self.assertIn("United Healthcare", synonym_response["answer"])

    def test_export_workbook_has_required_sheets(self) -> None:
        root = self.make_workspace()
        workbook_path = export_workbook(
            root,
            candidates=[
                {
                    "priority_rank": 1,
                    "payer_name": "United Healthcare",
                    "plan_name": "Commercial/EPO/PPO",
                    "service": "MRI brain without contrast",
                    "code": "70551",
                    "payer_rate": 1725.0,
                    "peer_median": 1510.0,
                    "variance_percent": 0.142,
                    "reason_flagged": "Potential pricing outlier.",
                    "confidence": 0.88,
                    "recommended_action": "Validate before outreach.",
                    "source_evidence": "sample://duke | row_4",
                }
            ],
            payer_summary=[
                {
                    "payer_name": "United Healthcare",
                    "number_of_plans": 1,
                    "number_of_services_analyzed": 1,
                    "number_of_outlier_flags": 1,
                    "average_variance": 0.142,
                    "top_concern": "MRI brain without contrast",
                    "confidence": 0.88,
                }
            ],
            source_manifest=[
                {
                    "source_name": "sample",
                    "source_type": "synthetic_fixture",
                    "source_url_or_path": "sample://duke",
                    "accessed_or_ingested_date": "2026-04-25",
                    "notes": "fixture",
                }
            ],
            methodology_lines=[
                "data sources: sample",
                "comparison method: peer median",
                "threshold: 20%",
                "limitations: sample fixture",
            ],
        )
        wb = load_workbook(workbook_path)
        self.assertEqual(
            wb.sheetnames,
            ["Outreach Candidates", "Payer Summary", "Methodology", "Source Manifest"],
        )
        headers = [cell.value for cell in wb["Outreach Candidates"][1]]
        self.assertIn("payer_name", headers)
        self.assertIn("source_evidence", headers)


if __name__ == "__main__":
    unittest.main()
