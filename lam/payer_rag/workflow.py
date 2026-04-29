from __future__ import annotations

import json
import re
from dataclasses import asdict, dataclass, field
from datetime import datetime
from pathlib import Path
from typing import Any, Iterable

from openpyxl import load_workbook

from .analyze import analyze_outliers, read_csv_rows
from .export import (
    build_validation_queue_rows,
    export_workbook,
    write_csv,
    write_real_data_acquisition_checklist,
    write_dashboard_html,
    write_summary_report,
    write_validation_checklist,
    write_validation_queue_csv,
)
from .ingest import DEFAULT_SERVICE_KEYWORDS, ingest_sources, load_source_manifest, write_ingestion_outputs
from .rag import ask_question, build_index
from lam.operator_platform.validators import (
    ArtifactContaminationValidator,
    FinalOutputGate,
    FinalOutputGateResult,
    GeographyValidator,
    ServiceScopeValidator,
    SourceRelevanceValidator,
)

RUNS_ROOT = Path("data/payer_rag_runs")
DEFAULT_DOMAIN = "insurance payer/plan pricing analysis"
MAJOR_PARAMETER_KEYS = {
    "geography",
    "state",
    "market",
    "domain",
    "requested_outputs",
    "source_constraints",
    "timeframe",
    "required_artifacts",
    "stakeholder_audience",
}
STATE_TOKEN_MAP = {
    "NC": ["nc", "north_carolina", "north carolina"],
    "VA": ["va", "virginia"],
}


@dataclass(slots=True)
class CurrentTaskContract:
    geography: str
    state: str
    market: str
    geography_explicit: bool = False
    domain: str = DEFAULT_DOMAIN
    service_focus: str = ""
    requested_outputs: list[str] = field(default_factory=list)
    source_constraints: list[str] = field(default_factory=list)
    timeframe: str = "current public data"
    required_artifacts: list[str] = field(default_factory=list)
    stakeholder_audience: str = "payer contracting and healthcare analytics stakeholders"
    invalid_geography_terms: list[str] = field(default_factory=list)
    invalid_source_terms: list[str] = field(default_factory=list)
    allowed_service_terms: list[str] = field(default_factory=list)
    disallowed_service_terms: list[str] = field(default_factory=list)

    def slug(self) -> str:
        return _slugify(self.geography)

    def label(self) -> str:
        return self.geography

    def prefix(self) -> str:
        return self.slug()


def _slugify(value: str) -> str:
    return re.sub(r"[^a-z0-9]+", "_", (value or "").strip().lower()).strip("_") or "unknown_market"


def _normalize_service_keywords(service_keywords: Iterable[str] | None) -> list[str]:
    values = [str(item).strip().lower() for item in (service_keywords or DEFAULT_SERVICE_KEYWORDS) if str(item).strip()]
    return list(dict.fromkeys(values))


def _infer_market_tokens(instruction: str) -> tuple[str, str, str, bool]:
    text = re.sub(r"\s+", " ", instruction or "").strip()
    patterns = [
        r"\bfor\s+([A-Z][A-Za-z .'-]+,\s*[A-Z]{2})\b",
        r"\bin\s+([A-Z][A-Za-z .'-]+,\s*[A-Z]{2})\b",
        r"\b([A-Z][A-Za-z .'-]+,\s*[A-Z]{2})\b",
    ]
    for pattern in patterns:
        match = re.search(pattern, text)
        if match:
            geography = re.sub(r"\s+", " ", match.group(1)).strip(" .")
            geography = re.sub(
                r"^(?:build|create|generate|run|review|prepare|produce|analyze|analyse|use)\s+(?:a|an|the)?\s*",
                "",
                geography,
                flags=re.IGNORECASE,
            ).strip(" .")
            parts = [part.strip() for part in geography.split(",", 1)]
            market = parts[0]
            state = parts[1] if len(parts) > 1 else ""
            return geography, market, state, True
    low = text.lower()
    if "fairfax" in low and "va" in low:
        return "Fairfax, VA", "Fairfax", "VA", True
    if "fairfax" in low:
        return "Fairfax, VA", "Fairfax", "VA", True
    if "durham" in low and "nc" in low:
        return "Durham, NC", "Durham", "NC", True
    if "durham" in low:
        return "Durham, NC", "Durham", "NC", True
    return "Durham, NC", "Durham", "NC", False


def extract_current_task_contract(instruction: str) -> CurrentTaskContract:
    geography, market, state, geography_explicit = _infer_market_tokens(instruction)
    low = (instruction or "").lower()
    requested_outputs: list[str] = []
    if any(token in low for token in ["rag", "vector store", "retriever", "index"]):
        requested_outputs.append("RAG/vector store")
    if any(token in low for token in ["summary", "stakeholder", "report"]):
        requested_outputs.append("stakeholder summary")
    if any(token in low for token in ["spreadsheet", "xlsx", "workbook", "outreach"]):
        requested_outputs.append("payer outreach spreadsheet")
    if "source manifest" in low or "sources" in low:
        requested_outputs.append("source manifest")
    if "data quality" in low:
        requested_outputs.append("data quality report")
    if "query examples" in low or "example queries" in low:
        requested_outputs.append("RAG query examples")
    if not requested_outputs:
        requested_outputs = [
            "RAG/vector store",
            "stakeholder summary",
            "payer outreach spreadsheet",
            "source manifest",
            "data quality report",
        ]
    source_constraints = [
        "public payer/plan/provider pricing information only",
        "no PHI",
        "no patient-level data",
        "no unsupported fairness claims",
    ]
    service_focus = "outpatient imaging" if all(token in low for token in ["outpatient", "imaging"]) or any(token in low for token in ["mri", "ct", "radiology", "ultrasound", "diagnostic imaging"]) else ""
    invalid_geography_terms = []
    invalid_source_terms = []
    if state.upper() == "VA" or "fairfax" in geography.lower():
        invalid_geography_terms = [
            "Duke University Hospital",
            "Duke Regional Hospital",
            "WakeMed",
            "UNC Health",
            "Blue Cross NC",
            "North Carolina",
            "Durham",
            "Raleigh",
            "Cary",
            "NC",
        ]
        invalid_source_terms = list(invalid_geography_terms)
    elif state.upper() == "NC" or "durham" in geography.lower():
        invalid_geography_terms = ["Fairfax", "Virginia", "VA"]
        invalid_source_terms = list(invalid_geography_terms)
    allowed_service_terms = []
    disallowed_service_terms = []
    if service_focus == "outpatient imaging":
        allowed_service_terms = [
            "mri",
            "magnetic resonance",
            "ct",
            "computed tomography",
            "cat scan",
            "x-ray",
            "xray",
            "radiology",
            "mammography",
            "ultrasound",
            "sonography",
            "imaging",
            "diagnostic imaging",
            "pet",
            "nuclear medicine",
        ]
        disallowed_service_terms = [
            "heart transplant",
            "transplant",
            "cabg",
            "vein harvest",
            "endoscopy",
            "colonoscopy",
            "thrombolysis",
            "pulmonary embolism",
            "contour defect",
            "surgery",
            "inpatient drg",
            "implant",
            "stoma",
        ]
    required_artifacts = [
        f"{_slugify(geography)}_payer_outreach_candidates.xlsx",
        f"{_slugify(geography)}_payer_dashboard.html",
        f"{_slugify(geography)}_summary_report.md",
        f"{_slugify(geography)}_source_manifest.csv",
        f"{_slugify(geography)}_data_quality_report.md",
        f"{_slugify(geography)}_contract_validation_queue.csv",
        f"{_slugify(geography)}_validation_checklist.md",
        "rag_index/payer_rag.db",
    ]
    return CurrentTaskContract(
        geography=geography,
        state=state,
        market=market,
        geography_explicit=geography_explicit,
        service_focus=service_focus,
        requested_outputs=requested_outputs,
        source_constraints=source_constraints,
        timeframe="current public data",
        required_artifacts=required_artifacts,
        invalid_geography_terms=invalid_geography_terms,
        invalid_source_terms=invalid_source_terms,
        allowed_service_terms=allowed_service_terms,
        disallowed_service_terms=disallowed_service_terms,
    )


def _runs_root() -> Path:
    RUNS_ROOT.mkdir(parents=True, exist_ok=True)
    return RUNS_ROOT


def _new_run_workspace(contract: CurrentTaskContract) -> Path:
    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    return _runs_root() / f"{contract.slug()}_{ts}"


def _task_contract_path(workspace: Path) -> Path:
    return workspace / "task_contract.json"


def _run_metadata_path(workspace: Path) -> Path:
    return workspace / "run_metadata.json"


def _artifact_paths(workspace: Path, contract: CurrentTaskContract) -> dict[str, str]:
    artifacts_root = workspace / "artifacts"
    normalized_root = workspace / "normalized"
    prefix = contract.prefix()
    return {
        "workbook_xlsx": str((artifacts_root / f"{prefix}_payer_outreach_candidates.xlsx").resolve()),
        "dashboard_html": str((artifacts_root / f"{prefix}_payer_dashboard.html").resolve()),
        "summary_report_md": str((artifacts_root / f"{prefix}_summary_report.md").resolve()),
        "source_manifest_csv": str((artifacts_root / f"{prefix}_source_manifest.csv").resolve()),
        "data_quality_report_md": str((artifacts_root / f"{prefix}_data_quality_report.md").resolve()),
        "outreach_candidates_csv": str((artifacts_root / f"{prefix}_outreach_candidates.csv").resolve()),
        "payer_summary_csv": str((artifacts_root / f"{prefix}_payer_summary.csv").resolve()),
        "validation_queue_csv": str((artifacts_root / f"{prefix}_contract_validation_queue.csv").resolve()),
        "validation_checklist_md": str((artifacts_root / f"{prefix}_validation_checklist.md").resolve()),
        "real_data_acquisition_checklist_md": str((artifacts_root / f"{prefix}_real_data_acquisition_checklist.md").resolve()),
        "geography_validation_report_md": str((artifacts_root / f"{prefix}_geography_validation.md").resolve()),
        "rag_index_db": str((artifacts_root / "rag_index" / "payer_rag.db").resolve()),
        "services_csv": str((normalized_root / "services.csv").resolve()),
        "rates_csv": str((normalized_root / "rates.csv").resolve()),
        "task_contract_json": str(_task_contract_path(workspace).resolve()),
        "primary_open_file": str((artifacts_root / f"{prefix}_payer_dashboard.html").resolve()),
    }


def _write_task_contract(workspace: Path, contract: CurrentTaskContract) -> Path:
    path = _task_contract_path(workspace)
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(json.dumps(asdict(contract), indent=2), encoding="utf-8")
    return path


def _write_run_metadata(
    workspace: Path,
    *,
    contract: CurrentTaskContract,
    invalidated_artifacts: list[str],
    reused_existing_outputs: bool,
    generation_ts: str,
) -> Path:
    payload = {
        "contract": asdict(contract),
        "generation_timestamp": generation_ts,
        "invalidated_artifacts": invalidated_artifacts,
        "reused_existing_outputs": reused_existing_outputs,
    }
    path = _run_metadata_path(workspace)
    path.write_text(json.dumps(payload, indent=2), encoding="utf-8")
    return path


def _read_task_contract(workspace: Path) -> CurrentTaskContract | None:
    path = _task_contract_path(workspace)
    if not path.exists():
        return None
    try:
        payload = json.loads(path.read_text(encoding="utf-8"))
        return CurrentTaskContract(**payload)
    except Exception:
        return None


def _workspaces_for_contract(contract: CurrentTaskContract) -> list[Path]:
    roots = [path for path in _runs_root().glob(f"{contract.slug()}_*") if path.is_dir()]
    roots.sort(key=lambda item: item.stat().st_mtime, reverse=True)
    return roots


def _all_workspaces() -> list[Path]:
    roots = [path for path in _runs_root().iterdir() if path.is_dir()]
    roots.sort(key=lambda item: item.stat().st_mtime, reverse=True)
    return roots


def _strict_contract_match(existing: CurrentTaskContract | None, contract: CurrentTaskContract) -> bool:
    if existing is None:
        return False
    for key in MAJOR_PARAMETER_KEYS:
        left = getattr(existing, key, None)
        right = getattr(contract, key, None)
        if left != right:
            return False
    return True


def _artifact_paths_from_existing(workspace: Path, contract: CurrentTaskContract) -> dict[str, str]:
    return _artifact_paths(workspace, contract)


def _stale_tokens(contract: CurrentTaskContract) -> list[str]:
    allowed = {_slugify(contract.geography), _slugify(contract.market), _slugify(contract.state)}
    for token in STATE_TOKEN_MAP.get(contract.state.upper(), []):
        allowed.add(token)
    candidates = ["durham", "durham_nc", "north_carolina", "nc", "fairfax", "fairfax_va", "virginia", "va"]
    return [token for token in candidates if token and token not in allowed]


def _contains_geo_token(text: str, token: str) -> bool:
    escaped = re.escape(token).replace("_", r"[\s_\-]+")
    pattern = rf"(?<![a-z0-9]){escaped}(?![a-z0-9])"
    return re.search(pattern, text.lower()) is not None


def _artifact_text_contains_stale_geography(path: Path, contract: CurrentTaskContract) -> tuple[bool, str]:
    try:
        if path.suffix.lower() == ".xlsx":
            wb = load_workbook(path, read_only=True, data_only=True)
            values: list[str] = []
            for ws in wb.worksheets[:4]:
                for row in ws.iter_rows(min_row=1, max_row=20, values_only=True):
                    values.extend(str(cell) for cell in row if cell not in {None, ""})
            text = " ".join(values).lower()
            wb.close()
        else:
            if path.suffix.lower() not in {".md", ".csv", ".html", ".json", ".txt"}:
                return False, ""
            text = path.read_text(encoding="utf-8", errors="ignore").lower()
    except Exception:
        return False, ""
    requested = contract.geography.lower()
    if requested not in text and path.suffix.lower() in {".md", ".html", ".xlsx"}:
        return True, f"missing requested geography marker '{contract.geography}' in {path.name}"
    for token in _stale_tokens(contract):
        if _contains_geo_token(text, token):
            return True, f"stale geography token '{token}' found in {path.name}"
    return False, ""


def _validate_workspace_relevance(workspace: Path, contract: CurrentTaskContract) -> tuple[bool, list[str]]:
    errors: list[str] = []
    existing = _read_task_contract(workspace)
    if not _strict_contract_match(existing, contract):
        errors.append("task contract mismatch")
        return False, errors
    artifacts = _artifact_paths_from_existing(workspace, contract)
    for key, raw_path in artifacts.items():
        path = Path(raw_path)
        if key == "primary_open_file":
            continue
        if not path.exists():
            errors.append(f"missing artifact: {path.name}")
            continue
        lower_name = path.name.lower()
        if any(token in lower_name for token in _stale_tokens(contract)):
            errors.append(f"stale filename token found: {path.name}")
        bad_text, reason = _artifact_text_contains_stale_geography(path, contract)
        if bad_text:
            errors.append(reason)
    return len(errors) == 0, errors


def _find_latest_valid_workspace(contract: CurrentTaskContract) -> tuple[Path | None, list[str]]:
    invalidated: list[str] = []
    for workspace in _all_workspaces():
        existing = _read_task_contract(workspace)
        if existing is not None and existing.geography != contract.geography:
            invalidated.append(f"{workspace.name}: geography mismatch ({existing.geography} != {contract.geography})")
            continue
        ok, errors = _validate_workspace_relevance(workspace, contract)
        if ok:
            return workspace, invalidated
        invalidated.append(f"{workspace.name}: {'; '.join(errors[:4])}")
    return None, invalidated


def _default_methodology_lines(contract: CurrentTaskContract, outlier_threshold: float) -> list[str]:
    return [
        f"geography: {contract.geography}",
        "data sources: Public provider transparency files, shoppable-service references, and payer transparency reference pages listed in the source manifest.",
        "comparison method: Compare negotiated rates within the same facility, setting, billing class, and service where public machine-readable rates are available.",
        f"threshold: Flag rates more than {outlier_threshold * 100:.0f}% above the peer median as potential pricing outliers.",
        "limitations: Public standard-charge data is not claim-level paid amount data; validate with contracts and sample claims before any outreach.",
        "interpretation guidance: Treat flags as review candidates or possible contract/pricing concerns, not legal conclusions.",
    ]


def _default_limitations(contract: CurrentTaskContract) -> list[str]:
    return [
        f"The build is scoped to {contract.geography} using public transparency and standard-charge data rather than PHI, claims, or remittance-level records.",
        "Some payer transparency and provider reference pages may need manual review if automated fetches are blocked.",
        "Shoppable-service catalogs improve consumer-friendly service matching but may not expose negotiated-rate rows for every payer-plan combination.",
        "Every flagged row should be validated with contracting terms and sample claims before payer outreach.",
    ]


def _synthetic_demo_source(contract: CurrentTaskContract) -> dict[str, Any]:
    focus = str(contract.service_focus or "outpatient imaging").strip() or "outpatient imaging"
    facility_name = f"{contract.market or 'Local'} Imaging Center"
    return {
        "source_name": f"Synthetic {contract.geography} {focus} demo dataset",
        "source_type": "synthetic_fixture",
        "source_url_or_path": f"sample://{contract.slug()}_{_slugify(focus)}_demo",
        "accessed_or_ingested_date": "",
        "geography": contract.geography,
        "notes": (
            f"Synthetic/demo {focus}-only dataset for {contract.geography}. "
            "Not validated public local evidence."
        ),
        "confidence": 0.35,
        "fixture_variant": "imaging_demo",
        "facility_name": facility_name,
    }


def _rename_output(path: Path, target_name: str) -> Path:
    target = path.with_name(target_name)
    if path.resolve() == target.resolve():
        return path
    if target.exists():
        target.unlink()
    path.replace(target)
    return target


def _rename_outputs_for_contract(workspace: Path, contract: CurrentTaskContract) -> None:
    prefix = contract.prefix()
    artifacts_root = workspace / "artifacts"
    mapping = {
        "durham_nc_payer_outreach_candidates.xlsx": f"{prefix}_payer_outreach_candidates.xlsx",
        "payer_dashboard.html": f"{prefix}_payer_dashboard.html",
        "summary_report.md": f"{prefix}_summary_report.md",
        "source_manifest.csv": f"{prefix}_source_manifest.csv",
        "data_quality_report.md": f"{prefix}_data_quality_report.md",
        "outreach_candidates.csv": f"{prefix}_outreach_candidates.csv",
        "payer_summary.csv": f"{prefix}_payer_summary.csv",
        "contract_validation_queue.csv": f"{prefix}_contract_validation_queue.csv",
        "validation_checklist.md": f"{prefix}_validation_checklist.md",
        "real_data_acquisition_checklist.md": f"{prefix}_real_data_acquisition_checklist.md",
    }
    for source_name, target_name in mapping.items():
        source = artifacts_root / source_name
        if source.exists():
            _rename_output(source, target_name)


def _write_geography_validation_report(
    workspace: Path,
    contract: CurrentTaskContract,
    artifacts: dict[str, str],
) -> tuple[Path, dict[str, Any]]:
    errors: list[str] = []
    checked: list[str] = []
    for key, raw_path in artifacts.items():
        if key in {"primary_open_file", "rag_index_db", "services_csv", "rates_csv", "task_contract_json", "geography_validation_report_md"}:
            continue
        path = Path(raw_path)
        if not path.exists():
            errors.append(f"missing artifact: {path.name}")
            continue
        checked.append(path.name)
        if any(_contains_geo_token(path.name.lower(), token) for token in _stale_tokens(contract)):
            errors.append(f"wrong geography in filename: {path.name}")
        bad_text, reason = _artifact_text_contains_stale_geography(path, contract)
        if bad_text:
            errors.append(reason)
    target = Path(artifacts["geography_validation_report_md"])
    status = "passed" if not errors else "failed"
    target.write_text(
        "# Geography Consistency Validation\n\n"
        f"- Requested geography: {contract.geography}\n"
        f"- Status: {status}\n"
        f"- Checked artifacts: {', '.join(checked) if checked else 'none'}\n\n"
        "## Findings\n\n"
        + ("\n".join(f"- {item}" for item in errors) if errors else "- All checked artifacts matched the current geography.") + "\n",
        encoding="utf-8",
    )
    return target, {"passed": not errors, "errors": errors, "checked_artifacts": checked}


def _required_artifacts_exist(artifacts: dict[str, str]) -> bool:
    required_keys = ["workbook_xlsx", "dashboard_html", "summary_report_md", "source_manifest_csv", "validation_queue_csv"]
    for key in required_keys:
        raw_path = str(artifacts.get(key, "")).strip()
        if not raw_path or not Path(raw_path).exists():
            return False
    return True


def _rebuild_payer_summary_from_candidates(candidates: list[dict]) -> list[dict]:
    rollup: dict[str, dict[str, Any]] = {}
    for row in candidates:
        payer_name = str(row.get("payer_name", "")).strip() or "unknown"
        current = rollup.setdefault(
            payer_name,
            {
                "payer_name": payer_name,
                "number_of_plans": 0,
                "plan_names": set(),
                "number_of_services_analyzed": 0,
                "number_of_outlier_flags": 0,
                "average_variance": 0.0,
                "top_concern": "",
                "confidence": 0.0,
                "_top_variance": -1.0,
            },
        )
        current["plan_names"].add(str(row.get("plan_name", "")).strip())
        current["number_of_services_analyzed"] += 1
        current["number_of_outlier_flags"] += 1
        variance = float(row.get("variance_percent", 0.0) or 0.0)
        current["average_variance"] += variance
        if variance > float(current["_top_variance"]):
            current["_top_variance"] = variance
            current["top_concern"] = str(row.get("service", "")).strip()
        current["confidence"] = max(float(current["confidence"]), float(row.get("confidence", 0.0) or 0.0))
    out: list[dict] = []
    for row in rollup.values():
        flags = int(row["number_of_outlier_flags"] or 0)
        out.append(
            {
                "payer_name": row["payer_name"],
                "number_of_plans": len(row["plan_names"]),
                "number_of_services_analyzed": row["number_of_services_analyzed"],
                "number_of_outlier_flags": flags,
                "average_variance": (float(row["average_variance"]) / flags) if flags else 0.0,
                "top_concern": row["top_concern"],
                "confidence": row["confidence"],
            }
        )
    out.sort(key=lambda item: (-int(item["number_of_outlier_flags"]), -float(item["average_variance"]), str(item["payer_name"])))
    return out


def _apply_scope_repair(
    *,
    root: Path,
    contract: CurrentTaskContract,
    artifacts: dict[str, str],
    candidates: list[dict],
    payer_summary: list[dict],
    source_manifest: list[dict],
    outlier_threshold: float,
) -> tuple[list[dict], list[dict], dict[str, Any]]:
    scope_validator = ServiceScopeValidator()
    filtered_candidates, removed_rows = scope_validator.filter_candidate_rows(contract=asdict(contract), candidate_rows=candidates)
    if not removed_rows:
        return candidates, payer_summary, {"attempted": False, "removed_count": 0}
    rebuilt_summary = _rebuild_payer_summary_from_candidates(filtered_candidates)
    write_csv(
        Path(artifacts["outreach_candidates_csv"]),
        filtered_candidates,
        [
            "priority_rank",
            "payer_name",
            "plan_name",
            "service",
            "code",
            "payer_rate",
            "peer_median",
            "peer_min",
            "peer_max",
            "variance_percent",
            "reason_flagged",
            "confidence",
            "recommended_action",
            "source_evidence",
            "compared_service",
            "facility_name",
        ],
    )
    write_csv(
        Path(artifacts["payer_summary_csv"]),
        rebuilt_summary,
        [
            "payer_name",
            "number_of_plans",
            "number_of_services_analyzed",
            "number_of_outlier_flags",
            "average_variance",
            "top_concern",
            "confidence",
        ],
    )
    validation_rows = build_validation_queue_rows(filtered_candidates)
    write_validation_queue_csv(root, validation_rows)
    write_validation_checklist(root, validation_rows, geography_label=contract.geography)
    synthetic_label = any(
        str(row.get("source_type", "")) == "synthetic_fixture" or str(row.get("source_url_or_path", "")).startswith("sample://")
        for row in source_manifest
    )
    write_real_data_acquisition_checklist(
        root,
        geography_label=contract.geography,
        service_focus=contract.service_focus or "outpatient imaging",
        synthetic_label=synthetic_label,
    )
    export_workbook(
        root,
        candidates=filtered_candidates,
        payer_summary=rebuilt_summary,
        source_manifest=source_manifest,
        methodology_lines=_default_methodology_lines(contract, outlier_threshold),
    )
    write_dashboard_html(root, geography_label=contract.geography)
    write_summary_report(
        root,
        candidates=filtered_candidates,
        payer_summary=rebuilt_summary,
        source_manifest=source_manifest,
        limitations=_default_limitations(contract),
        geography_label=contract.geography,
    )
    _rename_outputs_for_contract(root, contract)
    return filtered_candidates, rebuilt_summary, {"attempted": True, "removed_count": len(removed_rows)}


def _quarantine_artifacts(artifacts: dict[str, str], invalid_artifacts: list[str]) -> tuple[dict[str, str], dict[str, str]]:
    clean: dict[str, str] = {}
    quarantined: dict[str, str] = {}
    invalid = set(invalid_artifacts)
    for key, value in artifacts.items():
        if key in invalid:
            quarantined[key] = value
            continue
        clean[key] = value
    return clean, quarantined


def _run_validation_gate(
    *,
    root: Path,
    contract: CurrentTaskContract,
    artifacts: dict[str, str],
    candidates: list[dict],
    payer_summary: list[dict],
    source_manifest: list[dict],
    outlier_threshold: float,
    repair_state: dict[str, Any] | None = None,
) -> tuple[dict[str, Any], list[dict], list[dict], dict[str, str], dict[str, str]]:
    repair_state = dict(repair_state or {})
    contract_dict = asdict(contract)
    geography_result = GeographyValidator().validate(
        contract=contract_dict,
        artifacts=artifacts,
        source_rows=source_manifest,
        candidate_rows=candidates,
    )
    filtered_candidates = candidates
    filtered_summary = payer_summary
    service_result = ServiceScopeValidator().validate(
        contract=contract_dict,
        candidate_rows=filtered_candidates,
        artifact_paths=artifacts,
    )
    if not service_result.passed:
        filtered_candidates, filtered_summary, scope_repair = _apply_scope_repair(
            root=root,
            contract=contract,
            artifacts=artifacts,
            candidates=filtered_candidates,
            payer_summary=filtered_summary,
            source_manifest=source_manifest,
            outlier_threshold=outlier_threshold,
        )
        if scope_repair.get("attempted"):
            repair_state["service_scope_repair"] = scope_repair
            service_result = ServiceScopeValidator().validate(
                contract=contract_dict,
                candidate_rows=filtered_candidates,
                artifact_paths=artifacts,
            )
            service_result.repair_attempted = True
    source_result = SourceRelevanceValidator().validate(contract=contract_dict, source_rows=source_manifest)
    artifact_result = ArtifactContaminationValidator().validate(
        contract=contract_dict,
        artifacts=artifacts,
        geography_result=geography_result,
        service_result=service_result,
        source_result=source_result,
    )
    completion_passed = _required_artifacts_exist(artifacts)
    gate = FinalOutputGate().evaluate(
        validation_results=[geography_result, service_result, source_result, artifact_result],
        completion_passed=completion_passed,
        required_artifacts_exist=completion_passed,
    )
    clean_artifacts, quarantined_artifacts = _quarantine_artifacts(
        artifacts,
        list((artifact_result.metadata or {}).get("invalid_artifacts", []) or []),
    )
    validation_results = {
        "geography": geography_result.to_dict(),
        "service_scope": service_result.to_dict(),
        "source_relevance": source_result.to_dict(),
        "artifact_contamination": artifact_result.to_dict(),
    }
    return {
        "validation_results": validation_results,
        "final_output_gate": gate.to_dict(),
        "repair_state": repair_state,
        "quarantined_artifacts": quarantined_artifacts,
    }, filtered_candidates, filtered_summary, clean_artifacts, quarantined_artifacts


def build_workspace(
    *,
    contract: CurrentTaskContract,
    workspace: str | Path | None = None,
    manifest_path: str | Path | None = None,
    service_keywords: Iterable[str] | None = None,
    max_services_per_source: int = 24,
    outlier_threshold: float = 0.2,
    min_peer_count: int = 3,
    offline_fallback: bool = True,
    invalidated_artifacts: list[str] | None = None,
) -> dict[str, Any]:
    root = Path(workspace) if workspace else _new_run_workspace(contract)
    root.mkdir(parents=True, exist_ok=True)
    generation_ts = datetime.now().isoformat(timespec="seconds")
    manifest = load_source_manifest(manifest_path, offline=False, geography=contract.geography)
    pre_source_validation = SourceRelevanceValidator().validate(contract=asdict(contract), source_rows=manifest)
    allowed_source_rows = [
        dict(row)
        for row, evaluation in zip(manifest, list((pre_source_validation.metadata or {}).get("source_evaluations", []) or []))
        if bool(evaluation.get("allowed_as_evidence", False)) or bool(evaluation.get("synthetic", False))
    ]
    synthetic_only = bool((pre_source_validation.metadata or {}).get("synthetic_only"))
    if not allowed_source_rows:
        allowed_source_rows = [dict(row) for row in manifest if str(row.get("source_type", "")) == "synthetic_fixture"]
        synthetic_only = bool(allowed_source_rows)
    if not allowed_source_rows:
        allowed_source_rows = [_synthetic_demo_source(contract)]
        synthetic_only = True
    manifest_for_ingest = allowed_source_rows
    bundle = ingest_sources(
        manifest_for_ingest,
        service_keywords=_normalize_service_keywords(service_keywords),
        max_services_per_source=max_services_per_source,
        offline_fallback=offline_fallback,
    )
    outputs = write_ingestion_outputs(root, bundle, geography_label=contract.geography)
    rates = read_csv_rows(outputs["rates_csv"])
    plans = read_csv_rows(outputs["plans_csv"])
    payers = read_csv_rows(outputs["payers_csv"])
    services = read_csv_rows(outputs["services_csv"])
    analysis = analyze_outliers(
        rates=rates,
        plans=plans,
        payers=payers,
        services=services,
        outlier_threshold=outlier_threshold,
        min_peer_count=min_peer_count,
    )
    candidates = analysis["candidates"]
    payer_summary = analysis["payer_summary"]
    source_manifest = read_csv_rows(outputs["source_manifest_csv"])
    write_csv(
        root / "artifacts" / "outreach_candidates.csv",
        candidates,
        [
            "priority_rank",
            "payer_name",
            "plan_name",
            "service",
            "code",
            "payer_rate",
            "peer_median",
            "peer_min",
            "peer_max",
            "variance_percent",
            "reason_flagged",
            "confidence",
            "recommended_action",
            "source_evidence",
            "compared_service",
            "facility_name",
        ],
    )
    write_csv(
        root / "artifacts" / "payer_summary.csv",
        payer_summary,
        [
            "payer_name",
            "number_of_plans",
            "number_of_services_analyzed",
            "number_of_outlier_flags",
            "average_variance",
            "top_concern",
            "confidence",
        ],
    )
    validation_rows = build_validation_queue_rows(candidates)
    write_validation_queue_csv(root, validation_rows)
    write_validation_checklist(root, validation_rows, geography_label=contract.geography)
    write_real_data_acquisition_checklist(
        root,
        geography_label=contract.geography,
        service_focus=contract.service_focus or "outpatient imaging",
        synthetic_label=synthetic_only,
    )
    export_workbook(
        root,
        candidates=candidates,
        payer_summary=payer_summary,
        source_manifest=source_manifest,
        methodology_lines=_default_methodology_lines(contract, outlier_threshold),
        synthetic_label=synthetic_only,
    )
    write_dashboard_html(root, geography_label=contract.geography, synthetic_label=synthetic_only)
    write_summary_report(
        root,
        candidates=candidates,
        payer_summary=payer_summary,
        source_manifest=source_manifest,
        limitations=_default_limitations(contract),
        geography_label=contract.geography,
        synthetic_label=synthetic_only,
    )
    _rename_outputs_for_contract(root, contract)
    _write_task_contract(root, contract)
    _write_run_metadata(
        root,
        contract=contract,
        invalidated_artifacts=list(invalidated_artifacts or []),
        reused_existing_outputs=False,
        generation_ts=generation_ts,
    )
    build_index(root)
    artifacts = _artifact_paths(root, contract)
    outputs["source_manifest_csv"] = Path(artifacts["source_manifest_csv"])
    outputs["data_quality_report_md"] = Path(artifacts["data_quality_report_md"])
    validation_path, validation = _write_geography_validation_report(root, contract, artifacts)
    artifacts["geography_validation_report_md"] = str(validation_path.resolve())
    validation_bundle, candidates, payer_summary, clean_artifacts, quarantined_artifacts = _run_validation_gate(
        root=root,
        contract=contract,
        artifacts=artifacts,
        candidates=candidates,
        payer_summary=payer_summary,
        source_manifest=source_manifest,
        outlier_threshold=outlier_threshold,
        repair_state={"source_repair_attempted": not pre_source_validation.passed},
    )
    final_gate = dict(validation_bundle.get("final_output_gate", {}) or {})
    validation_results = dict(validation_bundle.get("validation_results", {}) or {})
    final_artifacts = clean_artifacts if not final_gate.get("passed", False) else artifacts
    completion_status = (
        "completed_demo_package"
        if synthetic_only and final_gate.get("passed", False)
        else ("partially_completed_validated" if final_gate.get("passed", False) else "blocked")
    )
    return {
        "workspace": str(root.resolve()),
        "artifact_paths": final_artifacts,
        "all_artifact_paths": artifacts,
        "ingestion_outputs": {key: str(Path(value).resolve()) for key, value in outputs.items()},
        "counts": {
            "payers": len(bundle["payers"]),
            "plans": len(bundle["plans"]),
            "services": len(bundle["services"]),
            "rates": len(bundle["rates"]),
            "outreach_candidates": len(candidates),
            "validation_queue": len(build_validation_queue_rows(candidates)),
        },
        "issues": list(bundle["issues"]),
        "top_candidates": candidates[:5],
        "current_task_contract": asdict(contract),
        "generation_timestamp": generation_ts,
        "invalidated_artifacts": list(invalidated_artifacts or []),
        "geography_validation": validation,
        "validation_results": validation_results,
        "final_output_gate": final_gate,
        "quarantined_artifacts": quarantined_artifacts,
        "repair_state": validation_bundle.get("repair_state", {}),
        "synthetic_only": synthetic_only,
        "completion_status": completion_status,
    }


def ensure_workspace(
    *,
    contract: CurrentTaskContract,
    manifest_path: str | Path | None = None,
    service_keywords: Iterable[str] | None = None,
    max_services_per_source: int = 24,
    outlier_threshold: float = 0.2,
    min_peer_count: int = 3,
    offline_fallback: bool = True,
    allow_reuse: bool = True,
) -> dict[str, Any]:
    reused_existing_outputs = False
    invalidated_artifacts: list[str] = []
    workspace: Path | None = None
    if not allow_reuse:
        invalidated_artifacts = [
            f"{path.name}: explicit geography request requires a fresh run"
            for path in _all_workspaces()[:10]
        ]
    if allow_reuse:
        workspace, invalidated_artifacts = _find_latest_valid_workspace(contract)
        reused_existing_outputs = workspace is not None
    if workspace is not None:
        root = workspace
        artifacts = _artifact_paths(root, contract)
        validation_path, validation = _write_geography_validation_report(root, contract, artifacts)
        artifacts["geography_validation_report_md"] = str(validation_path.resolve())
        counts = {
            "payers": len(read_csv_rows(root / "normalized" / "payers.csv")) if (root / "normalized" / "payers.csv").exists() else 0,
            "plans": len(read_csv_rows(root / "normalized" / "plans.csv")) if (root / "normalized" / "plans.csv").exists() else 0,
            "services": len(read_csv_rows(root / "normalized" / "services.csv")) if (root / "normalized" / "services.csv").exists() else 0,
            "rates": len(read_csv_rows(root / "normalized" / "rates.csv")) if (root / "normalized" / "rates.csv").exists() else 0,
            "outreach_candidates": len(read_csv_rows(Path(artifacts["outreach_candidates_csv"]))) if Path(artifacts["outreach_candidates_csv"]).exists() else 0,
            "validation_queue": len(read_csv_rows(Path(artifacts["validation_queue_csv"]))) if Path(artifacts["validation_queue_csv"]).exists() else 0,
        }
        all_candidates = read_csv_rows(Path(artifacts["outreach_candidates_csv"])) if Path(artifacts["outreach_candidates_csv"]).exists() else []
        payer_summary_rows = read_csv_rows(Path(artifacts["payer_summary_csv"])) if Path(artifacts["payer_summary_csv"]).exists() else []
        source_manifest_rows = read_csv_rows(Path(artifacts["source_manifest_csv"])) if Path(artifacts["source_manifest_csv"]).exists() else []
        validation_bundle, filtered_candidates, filtered_summary, clean_artifacts, quarantined_artifacts = _run_validation_gate(
            root=root,
            contract=contract,
            artifacts=artifacts,
            candidates=all_candidates,
            payer_summary=payer_summary_rows,
            source_manifest=source_manifest_rows,
            outlier_threshold=outlier_threshold,
            repair_state={},
        )
        return {
            "workspace": str(root.resolve()),
            "artifact_paths": clean_artifacts if not bool((validation_bundle.get("final_output_gate", {}) or {}).get("passed", False)) else artifacts,
            "all_artifact_paths": artifacts,
            "counts": counts,
            "top_candidates": filtered_candidates[:5],
            "current_task_contract": asdict(contract),
            "reused_existing_outputs": reused_existing_outputs,
            "invalidated_artifacts": invalidated_artifacts,
            "generation_timestamp": json.loads(_run_metadata_path(root).read_text(encoding="utf-8")).get("generation_timestamp", "")
            if _run_metadata_path(root).exists()
            else "",
            "geography_validation": validation,
            "validation_results": validation_bundle.get("validation_results", {}),
            "final_output_gate": validation_bundle.get("final_output_gate", {}),
            "quarantined_artifacts": quarantined_artifacts,
            "repair_state": validation_bundle.get("repair_state", {}),
            "synthetic_only": bool((validation_bundle.get("validation_results", {}).get("source_relevance", {}) or {}).get("metadata", {}).get("synthetic_only")),
            "completion_status": (
                "completed_demo_package"
                if bool((validation_bundle.get("validation_results", {}).get("source_relevance", {}) or {}).get("metadata", {}).get("synthetic_only"))
                and bool((validation_bundle.get("final_output_gate", {}) or {}).get("passed", False))
                else ("partially_completed_validated" if bool((validation_bundle.get("final_output_gate", {}) or {}).get("passed", False)) else "blocked")
            ),
        }
    return build_workspace(
        contract=contract,
        workspace=None,
        manifest_path=manifest_path,
        service_keywords=service_keywords,
        max_services_per_source=max_services_per_source,
        outlier_threshold=outlier_threshold,
        min_peer_count=min_peer_count,
        offline_fallback=offline_fallback,
        invalidated_artifacts=invalidated_artifacts,
    ) | {"reused_existing_outputs": False}


def ask_workspace_question(
    question: str,
    *,
    contract: CurrentTaskContract,
    workspace: str | Path | None = None,
) -> dict[str, Any]:
    root = Path(workspace) if workspace else (_find_latest_valid_workspace(contract)[0] or Path())
    if not root or not root.exists():
        return {
            "answer": f"No valid workspace exists yet for {contract.geography}. Rebuild the package for this geography first.",
            "sources": [],
            "workspace": "",
            "artifacts": {},
            "current_task_contract": asdict(contract),
        }
    response = ask_question(root, question)
    response["workspace"] = str(root.resolve())
    response["artifacts"] = _artifact_paths(root, contract)
    response["current_task_contract"] = asdict(contract)
    return response
