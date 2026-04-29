from __future__ import annotations

import csv
import html
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter

from .analyze import read_csv_rows


def write_csv(path: str | Path, rows: list[dict], fieldnames: list[str]) -> Path:
    target = Path(path)
    target.parent.mkdir(parents=True, exist_ok=True)
    with target.open("w", encoding="utf-8", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=fieldnames)
        writer.writeheader()
        for row in rows:
            writer.writerow({field: row.get(field, "") for field in fieldnames})
    return target


def build_validation_queue_rows(candidates: list[dict], max_rows: int = 25) -> list[dict]:
    rows: list[dict] = []
    for row in candidates[:max_rows]:
        rows.append(
            {
                "priority_rank": row.get("priority_rank", ""),
                "payer_name": row.get("payer_name", ""),
                "plan_name": row.get("plan_name", ""),
                "service": row.get("service", ""),
                "code": row.get("code", ""),
                "payer_rate": row.get("payer_rate", ""),
                "peer_median": row.get("peer_median", ""),
                "variance_percent": row.get("variance_percent", ""),
                "source_evidence": row.get("source_evidence", ""),
                "validation_status": "pending_contract_review",
                "contracting_owner": "",
                "sample_claim_needed": "yes",
                "sample_claim_status": "not_started",
                "recommended_validation_step": (
                    "Confirm contract terms, review sample claims, and verify service mapping before any outreach."
                ),
                "notes": "",
            }
        )
    return rows


def write_validation_queue_csv(workspace: str | Path, rows: list[dict]) -> Path:
    return write_csv(
        Path(workspace) / "artifacts" / "contract_validation_queue.csv",
        rows,
        [
            "priority_rank",
            "payer_name",
            "plan_name",
            "service",
            "code",
            "payer_rate",
            "peer_median",
            "variance_percent",
            "source_evidence",
            "validation_status",
            "contracting_owner",
            "sample_claim_needed",
            "sample_claim_status",
            "recommended_validation_step",
            "notes",
        ],
    )


def write_validation_checklist(workspace: str | Path, rows: list[dict], *, geography_label: str = "Durham, NC") -> Path:
    target = Path(workspace) / "artifacts" / "validation_checklist.md"
    top_lines = "\n".join(
        f"- {row['payer_name']} / {row['plan_name']} / {row['service']} ({float(row['variance_percent']) * 100:.1f}% above peer median)"
        for row in rows[:10]
        if row.get("variance_percent") not in {"", None}
    ) or "- No validation candidates were queued."
    target.write_text(
        "# Contract Validation Checklist\n\n"
        f"- Requested geography: {geography_label}\n\n"
        "Use this list before any payer outreach.\n\n"
        "## Required checks\n\n"
        "- Confirm the service mapping between the transparency file and the operational service line.\n"
        "- Validate payer and plan naming against current contract metadata.\n"
        "- Pull sample claims or remittance examples for the flagged service.\n"
        "- Confirm whether billed, negotiated, and allowed amounts are being compared appropriately.\n"
        "- Document any known carve-outs, stop-loss terms, or case-rate provisions.\n"
        "- Escalate only after contracting and revenue-cycle review confirms the concern.\n\n"
        "## Top queued rows\n\n"
        f"{top_lines}\n",
        encoding="utf-8",
    )
    return target


def write_real_data_acquisition_checklist(
    workspace: str | Path,
    *,
    geography_label: str = "Fairfax, VA",
    service_focus: str = "outpatient imaging",
    synthetic_label: bool = False,
) -> Path:
    target = Path(workspace) / "artifacts" / "real_data_acquisition_checklist.md"
    parts = [item.strip() for item in str(geography_label or "").split(",", 1)]
    region_name = parts[0] if parts else geography_label
    state_name = parts[1] if len(parts) > 1 else ""
    synthetic_note = (
        "A synthetic/demo package was generated because validated local public evidence was unavailable in this run.\n\n"
        if synthetic_label
        else ""
    )
    target.write_text(
        "# Real Data Acquisition Checklist\n\n"
        f"- Requested geography: {geography_label}\n"
        f"- Requested service scope: {service_focus}\n\n"
        f"{synthetic_note}"
        "## Collect real evidence\n\n"
        f"- Identify {geography_label} or broader {state_name} outpatient imaging provider transparency files.\n"
        "- Add locally relevant standard-charge or shoppable-service sources to the source manifest.\n"
        "- Prefer provider or payer files that explicitly identify MRI, CT, ultrasound, mammography, x-ray, or diagnostic imaging services.\n"
        "- Exclude stale or out-of-market artifacts from final evidence.\n"
        f"- Verify any national payer transparency source is filtered to {region_name} before using it as evidence.\n\n"
        "## Validate before outreach\n\n"
        "- Confirm service mapping for each imaging row with contracting or revenue-cycle stakeholders.\n"
        "- Pull sample claims or remittance examples for the top flagged outpatient imaging services.\n"
        "- Validate payer and plan naming against current contract metadata.\n"
        "- Confirm whether billed, negotiated, and allowed amounts are comparable for the flagged service.\n"
        "- Keep synthetic/demo results separate from validated public local evidence.\n",
        encoding="utf-8",
    )
    return target


def _format_sheet(ws) -> None:
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    for cell in ws[1]:
        cell.font = Font(bold=True)
    for idx, column_cells in enumerate(ws.columns, start=1):
        width = max(len(str(cell.value or "")) for cell in column_cells[:50]) + 2
        ws.column_dimensions[get_column_letter(idx)].width = min(max(width, 14), 48)


def export_workbook(
    workspace: str | Path,
    *,
    candidates: list[dict],
    payer_summary: list[dict],
    source_manifest: list[dict],
    methodology_lines: list[str],
    synthetic_label: bool = False,
) -> Path:
    root = Path(workspace)
    target = root / "artifacts" / "durham_nc_payer_outreach_candidates.xlsx"
    wb = Workbook()

    ws = wb.active
    ws.title = "Outreach Candidates"
    candidate_columns = [
        "priority_rank",
        "payer_name",
        "plan_name",
        "service",
        "code",
        "payer_rate",
        "peer_median",
        "variance_percent",
        "reason_flagged",
        "confidence",
        "recommended_action",
        "source_evidence",
    ]
    ws.append(candidate_columns)
    for row in candidates:
        ws.append([row.get(column, "") for column in candidate_columns])
    for cell in ws["H"][1:]:
        cell.number_format = "0.0%"
    for cell in ws["F"][1:] + ws["G"][1:]:
        cell.number_format = "$#,##0.00"
    _format_sheet(ws)

    summary_ws = wb.create_sheet("Payer Summary")
    summary_columns = [
        "payer_name",
        "number_of_plans",
        "number_of_services_analyzed",
        "number_of_outlier_flags",
        "average_variance",
        "top_concern",
        "confidence",
    ]
    summary_ws.append(summary_columns)
    for row in payer_summary:
        summary_ws.append([row.get(column, "") for column in summary_columns])
    for cell in summary_ws["E"][1:]:
        cell.number_format = "0.0%"
    _format_sheet(summary_ws)

    methodology_ws = wb.create_sheet("Methodology")
    methodology_ws.append(["section", "details"])
    if synthetic_label:
        methodology_ws.append(["synthetic_label", "This workbook uses synthetic/demo data and is not public Fairfax evidence."])
    for line in methodology_lines:
        if ":" in line:
            section, details = line.split(":", 1)
        else:
            section, details = "note", line
        methodology_ws.append([section.strip(), details.strip()])
    _format_sheet(methodology_ws)

    manifest_ws = wb.create_sheet("Source Manifest")
    manifest_columns = ["source_name", "source_type", "source_url_or_path", "accessed_or_ingested_date", "notes"]
    manifest_ws.append(manifest_columns)
    for row in source_manifest:
        manifest_ws.append([row.get(column, "") for column in manifest_columns])
    _format_sheet(manifest_ws)

    target.parent.mkdir(parents=True, exist_ok=True)
    wb.save(target)
    return target


def write_summary_report(
    workspace: str | Path,
    *,
    candidates: list[dict],
    payer_summary: list[dict],
    source_manifest: list[dict],
    limitations: list[str],
    geography_label: str = "Durham, NC",
    synthetic_label: bool = False,
) -> Path:
    root = Path(workspace)
    top_candidates = "\n".join(
        f"- {row['payer_name']} / {row['plan_name']} / {row['service']} "
        f"({float(row['variance_percent']) * 100:.1f}% above peer median)"
        for row in candidates[:10]
    ) or "- No candidates were flagged."
    manifest_lines = "\n".join(
        f"- {row['source_name']}: {row['source_url_or_path']}"
        for row in source_manifest
    )
    payer_lines = "\n".join(
        f"- {row['payer_name']}: {row['number_of_outlier_flags']} flags, top concern {row['top_concern'] or 'none'}"
        for row in payer_summary[:10]
    ) or "- No payer summaries available."
    limitation_lines = "\n".join(f"- {line}" for line in limitations)
    target = root / "artifacts" / "summary_report.md"
    target.write_text(
        f"# {geography_label} Payer Pricing Review\n\n"
        + ("**Synthetic/demo build:** This package does not represent validated public local pricing evidence.\n\n" if synthetic_label else "")
        +
        "## Summary\n\n"
        f"This review uses public {geography_label} provider transparency data, payer reference sources, and a local retriever over normalized payer, plan, "
        "service, and pricing records. Flagged items are potential pricing outliers that warrant validation before outreach.\n\n"
        "## Top Outreach Candidates\n\n"
        f"{top_candidates}\n\n"
        "## Payer Summary\n\n"
        f"{payer_lines}\n\n"
        "## Sources\n\n"
        f"{manifest_lines}\n\n"
        "## Limitations\n\n"
        f"{limitation_lines}\n\n"
        "## Validation Before Outreach\n\n"
        "- Review `contract_validation_queue.csv` with contracting and revenue-cycle stakeholders.\n"
        "- Validate the top rows with sample claims before contacting any payer.\n",
        encoding="utf-8",
    )
    return target


def write_dashboard_html(workspace: str | Path, *, geography_label: str = "Durham, NC", synthetic_label: bool = False) -> Path:
    root = Path(workspace)
    candidates = read_csv_rows(root / "artifacts" / "outreach_candidates.csv")
    payer_summary = read_csv_rows(root / "artifacts" / "payer_summary.csv")
    manifest = read_csv_rows(root / "artifacts" / "source_manifest.csv")
    top_rows = "".join(
        "<tr>"
        f"<td>{row['priority_rank']}</td>"
        f"<td>{html.escape(row['payer_name'])}</td>"
        f"<td>{html.escape(row['plan_name'])}</td>"
        f"<td>{html.escape(row['service'])}</td>"
        f"<td>{float(row['payer_rate']):,.2f}</td>"
        f"<td>{float(row['peer_median']):,.2f}</td>"
        f"<td>{float(row['variance_percent']) * 100:.1f}%</td>"
        "</tr>"
        for row in candidates[:15]
    )
    payer_rows = "".join(
        "<tr>"
        f"<td>{html.escape(row['payer_name'])}</td>"
        f"<td>{row['number_of_outlier_flags']}</td>"
        f"<td>{float(row['average_variance']) * 100:.1f}%</td>"
        f"<td>{html.escape(row['top_concern'])}</td>"
        "</tr>"
        for row in payer_summary[:10]
    )
    source_rows = "".join(
        "<li><a href=\"{url}\">{name}</a></li>".format(
            url=html.escape(row["source_url_or_path"]),
            name=html.escape(row["source_name"]),
        )
        for row in manifest
    )
    target = root / "artifacts" / "shopping_dashboard_placeholder.html"
    target = root / "artifacts" / "payer_dashboard.html"
    synthetic_badge = "<span class='pill' style='background:#fee2e2;color:#991b1b;margin-left:8px;'>Synthetic / demo only</span>" if synthetic_label else ""
    target.write_text(
        f"<!doctype html><html><head><meta charset='utf-8'><title>{html.escape(geography_label)} Payer Review</title>"
        "<style>body{font-family:Georgia,serif;background:#f5f7fb;color:#182230;padding:32px} "
        ".grid{display:grid;grid-template-columns:1.5fr 1fr;gap:24px} .card{background:#fff;border:1px solid #dde5f0;"
        "border-radius:18px;padding:20px;box-shadow:0 10px 30px rgba(18,35,56,.06)} table{width:100%;border-collapse:collapse}"
        "th,td{padding:10px;border-bottom:1px solid #eef2f7;text-align:left} h1,h2{margin-top:0} .pill{display:inline-block;"
        "padding:6px 10px;border-radius:999px;background:#d8ebff;color:#15406b;font-size:12px;font-weight:700}</style></head><body>"
        f"<span class='pill'>Potential pricing outliers only</span>{synthetic_badge}<h1>{html.escape(geography_label)} Payer Pricing Review</h1>"
        f"<p>{'This is a synthetic/demo fallback and is not stakeholder-ready public local evidence. ' if synthetic_label else ''}This dashboard summarizes public {html.escape(geography_label)} pricing data and highlights payer-plan combinations that appear higher than local peers for comparable services. All items require contract validation before outreach.</p>"
        "<div class='grid'><div class='card'><h2>Top Outreach Candidates</h2><table><thead><tr><th>Rank</th><th>Payer</th><th>Plan</th><th>Service</th><th>Payer Rate</th><th>Peer Median</th><th>Variance</th></tr></thead><tbody>"
        f"{top_rows}</tbody></table></div><div class='card'><h2>Payer Summary</h2><table><thead><tr><th>Payer</th><th>Flags</th><th>Avg Variance</th><th>Top Concern</th></tr></thead><tbody>{payer_rows}</tbody></table><h2 style='margin-top:24px'>Sources</h2><ul>{source_rows}</ul></div></div></body></html>",
        encoding="utf-8",
    )
    return target
