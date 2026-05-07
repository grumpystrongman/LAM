from __future__ import annotations

import csv
import json
import re
from dataclasses import asdict, dataclass, field
from pathlib import Path
from typing import Any, Dict, Iterable, List

from openpyxl import load_workbook


@dataclass(slots=True)
class ValidationViolation:
    artifact: str
    matched_text: str
    reason: str

    def to_dict(self) -> Dict[str, Any]:
        return asdict(self)


@dataclass(slots=True)
class ValidationResult:
    validator: str
    passed: bool
    severity: str = "blocking"
    violations: List[ValidationViolation] = field(default_factory=list)
    repair_attempted: bool = False
    final_status: str = ""
    score: float = 1.0
    metadata: Dict[str, Any] = field(default_factory=dict)

    def to_dict(self) -> Dict[str, Any]:
        payload = asdict(self)
        payload["issue_count"] = len(self.violations)
        payload["final_status"] = self.final_status or ("passed" if self.passed else "failed")
        return payload


@dataclass(slots=True)
class FinalOutputGateResult:
    passed: bool
    severity: str
    blocking_failures: List[str]
    required_repairs: List[str]
    issue_count: int

    def to_dict(self) -> Dict[str, Any]:
        return asdict(self)


IMAGING_ALLOWED_TERMS = [
    "mri",
    "magnetic resonance",
    "ct",
    "computed tomography",
    "cat scan",
    "x-ray",
    "xray",
    "radiology",
    "mammography",
    "ultrasound",
    "sonography",
    "imaging",
    "diagnostic imaging",
    "pet",
    "nuclear medicine",
]

IMAGING_DISALLOWED_TERMS = [
    "heart transplant",
    "transplant",
    "cabg",
    "vein harvest",
    "endoscopy",
    "colonoscopy",
    "thrombolysis",
    "pulmonary embolism",
    "contour defect",
    "surgery",
    "inpatient drg",
    "implant",
    "stoma",
]

DEFAULT_INVALID_GEOGRAPHY_TERMS = {
    "Fairfax, VA": [
        "Duke University Hospital",
        "Duke Regional Hospital",
        "WakeMed",
        "UNC Health",
        "Blue Cross NC",
        "North Carolina",
        "Durham",
        "Raleigh",
        "Cary",
        "NC",
    ],
    "Durham, NC": [
        "Fairfax",
        "Virginia",
        "VA",
    ],
}


def _term_in_text(text: str, term: str) -> bool:
    escaped = re.escape(str(term or "").strip().lower())
    if not escaped:
        return False
    sep_pattern = r"[\s_\-]+"
    pattern = rf"(?<![a-z0-9]){escaped.replace('_', sep_pattern)}(?![a-z0-9])"
    return re.search(pattern, str(text or "").lower()) is not None


def _read_textish_artifact(path: Path) -> str:
    if not path.exists():
        return ""
    suffix = path.suffix.lower()
    try:
        if suffix == ".xlsx":
            wb = load_workbook(path, read_only=True, data_only=True)
            values: List[str] = []
            for ws in wb.worksheets[:6]:
                for row in ws.iter_rows(min_row=1, max_row=80, values_only=True):
                    values.extend(str(cell) for cell in row if cell not in {None, ""})
            wb.close()
            return " ".join(values)
        if suffix == ".csv":
            rows: List[str] = []
            with path.open("r", encoding="utf-8", errors="ignore", newline="") as handle:
                reader = csv.reader(handle)
                for idx, row in enumerate(reader):
                    rows.extend(str(cell) for cell in row if str(cell).strip())
                    if idx >= 80:
                        break
            return " ".join(rows)
        if suffix in {".md", ".txt", ".html", ".json"}:
            return path.read_text(encoding="utf-8", errors="ignore")
    except Exception:
        return ""
    return ""


def _geography_reflected_in_probe(probe: str, geography: str) -> bool:
    geography_value = str(geography or "").strip()
    if not geography_value:
        return True
    probe_low = str(probe or "").lower()
    geography_low = geography_value.lower()
    if geography_low in probe_low or _term_in_text(probe_low, geography_low):
        return True
    slug = geography_low.replace(", ", "_").replace(" ", "_").replace("-", "_")
    dash_slug = geography_low.replace(", ", "-").replace(" ", "-")
    compact = geography_low.replace(",", "").replace(" ", "")
    return any(token and token in probe_low for token in {slug, dash_slug, compact})


def _contract_invalid_terms(contract: Dict[str, Any]) -> List[str]:
    geography = str(contract.get("geography", "")).strip()
    explicit = [str(item).strip() for item in (contract.get("invalid_geography_terms") or contract.get("invalid_geographies") or []) if str(item).strip()]
    seeded = list(DEFAULT_INVALID_GEOGRAPHY_TERMS.get(geography, []))
    out: List[str] = []
    for item in seeded + explicit:
        if item and item not in out:
            out.append(item)
    return out


def _allowed_service_terms(contract: Dict[str, Any]) -> List[str]:
    explicit = [str(item).strip().lower() for item in (contract.get("allowed_service_terms") or []) if str(item).strip()]
    focus = str(contract.get("service_focus", "")).lower()
    if explicit:
        return explicit
    if "imaging" in focus:
        return list(IMAGING_ALLOWED_TERMS)
    return []


def _disallowed_service_terms(contract: Dict[str, Any]) -> List[str]:
    explicit = [str(item).strip().lower() for item in (contract.get("disallowed_service_terms") or []) if str(item).strip()]
    focus = str(contract.get("service_focus", "")).lower()
    if explicit:
        return explicit
    if "imaging" in focus:
        return list(IMAGING_DISALLOWED_TERMS)
    return []


class GeographyValidator:
    def validate(
        self,
        *,
        contract: Dict[str, Any],
        artifacts: Dict[str, str],
        source_rows: Iterable[Dict[str, Any]] | None = None,
        candidate_rows: Iterable[Dict[str, Any]] | None = None,
        final_text: str = "",
    ) -> ValidationResult:
        invalid_terms = _contract_invalid_terms(contract)
        violations: List[ValidationViolation] = []
        ignored_artifacts = {"task_contract_json", "geography_validation_report_md"}
        for key, raw_path in (artifacts or {}).items():
            if key in ignored_artifacts:
                continue
            path = Path(str(raw_path))
            probe = f"{path.name} {str(path)} {_read_textish_artifact(path)}"
            for term in invalid_terms:
                if _term_in_text(probe, term):
                    violations.append(
                        ValidationViolation(
                            artifact=key,
                            matched_text=term,
                            reason=f"Out-of-geography source or text for {contract.get('geography', '')} task",
                        )
                    )
        for row in list(source_rows or []):
            probe = " ".join(
                [
                    str(row.get("source_name", "")),
                    str(row.get("geography", "")),
                    str(row.get("notes", "")),
                    str(row.get("source_url_or_path", "")),
                ]
            )
            for term in invalid_terms:
                if _term_in_text(probe, term):
                    violations.append(
                        ValidationViolation(
                            artifact=str(row.get("source_name", "source_manifest")),
                            matched_text=term,
                            reason=f"Out-of-geography source for {contract.get('geography', '')} task",
                        )
                    )
        for row in list(candidate_rows or []):
            probe = " ".join(
                [
                    str(row.get("service", "")),
                    str(row.get("facility_name", "")),
                    str(row.get("source_evidence", "")),
                    str(row.get("payer_name", "")),
                ]
            )
            for term in invalid_terms:
                if _term_in_text(probe, term):
                    violations.append(
                        ValidationViolation(
                            artifact=str(row.get("service", "candidate")),
                            matched_text=term,
                            reason=f"Out-of-geography evidence in stakeholder candidate rows for {contract.get('geography', '')}",
                        )
                    )
        if final_text:
            for term in invalid_terms:
                if _term_in_text(final_text, term):
                    violations.append(
                        ValidationViolation(
                            artifact="final_response",
                            matched_text=term,
                            reason=f"Final response references conflicting geography for {contract.get('geography', '')}",
                        )
                    )
        return ValidationResult(
            validator="GeographyValidator",
            passed=not violations,
            severity="blocking",
            violations=violations[:80],
            score=1.0 if not violations else 0.0,
        )


class ServiceScopeValidator:
    def validate(
        self,
        *,
        contract: Dict[str, Any],
        candidate_rows: Iterable[Dict[str, Any]] | None = None,
        artifact_paths: Dict[str, str] | None = None,
        final_text: str = "",
    ) -> ValidationResult:
        allowed = _allowed_service_terms(contract)
        disallowed = _disallowed_service_terms(contract)
        if not allowed and not disallowed:
            return ValidationResult(validator="ServiceScopeValidator", passed=True, severity="blocking", score=1.0)
        violations: List[ValidationViolation] = []
        ignored_artifacts = {"task_contract_json", "geography_validation_report_md"}
        for row in list(candidate_rows or []):
            service = str(row.get("service", "")).strip()
            service_low = service.lower()
            for term in disallowed:
                if term in service_low:
                    violations.append(
                        ValidationViolation(
                            artifact=str(row.get("priority_rank", "candidate")),
                            matched_text=term,
                            reason=f"Out-of-scope service found for {contract.get('service_focus', 'requested scope')}",
                        )
                    )
                    break
        for key, raw_path in (artifact_paths or {}).items():
            if key in ignored_artifacts:
                continue
            text = _read_textish_artifact(Path(str(raw_path)))
            for term in disallowed:
                if term and term in text.lower():
                    violations.append(
                        ValidationViolation(
                            artifact=key,
                            matched_text=term,
                            reason=f"Artifact contains service-line contamination outside {contract.get('service_focus', 'requested scope')}",
                        )
                    )
        if final_text:
            for term in disallowed:
                if term and term in final_text.lower():
                    violations.append(
                        ValidationViolation(
                            artifact="final_response",
                            matched_text=term,
                            reason=f"Final response mentions out-of-scope service for {contract.get('service_focus', 'requested scope')}",
                        )
                    )
        return ValidationResult(
            validator="ServiceScopeValidator",
            passed=not violations,
            severity="blocking",
            violations=violations[:80],
            score=1.0 if not violations else 0.0,
            metadata={"service_focus": contract.get("service_focus", "")},
        )

    def filter_candidate_rows(self, *, contract: Dict[str, Any], candidate_rows: Iterable[Dict[str, Any]]) -> tuple[List[Dict[str, Any]], List[Dict[str, Any]]]:
        disallowed = _disallowed_service_terms(contract)
        if not disallowed:
            return [dict(row) for row in candidate_rows], []
        kept: List[Dict[str, Any]] = []
        removed: List[Dict[str, Any]] = []
        for row in candidate_rows:
            service_low = str(row.get("service", "")).lower()
            if any(term in service_low for term in disallowed):
                removed.append(dict(row))
                continue
            kept.append(dict(row))
        return kept, removed


class SourceRelevanceValidator:
    def validate(self, *, contract: Dict[str, Any], source_rows: Iterable[Dict[str, Any]] | None = None) -> ValidationResult:
        geography = str(contract.get("geography", "")).strip().lower()
        state = str(contract.get("state", "")).strip().lower()
        service_focus = str(contract.get("service_focus", "")).strip().lower()
        invalid_terms = _contract_invalid_terms(contract)
        evaluations: List[Dict[str, Any]] = []
        violations: List[ValidationViolation] = []
        allowed_evidence_count = 0
        for row in list(source_rows or []):
            source_name = str(row.get("source_name", "")).strip()
            source_type = str(row.get("source_type", "")).strip()
            source_geo = str(row.get("geography", "")).strip().lower()
            notes = str(row.get("notes", "")).strip().lower()
            url = str(row.get("source_url_or_path", "")).strip()
            probe = " ".join([source_name, source_geo, notes, url]).lower()
            geography_match = bool(geography and geography in probe) or bool(state and state in probe) or source_geo in {"national", "synthetic"}
            service_scope_match = True
            if service_focus and "imaging" in service_focus:
                service_scope_match = any(term in notes or term in source_name.lower() or term in url.lower() for term in ["mri", "ct", "imaging", "radiology", "x-ray", "ultrasound", "shoppable"])
            synthetic = source_type == "synthetic_fixture" or url.startswith("sample://")
            blocked_geo = next((term for term in invalid_terms if _term_in_text(probe, term)), "")
            allowed_as_evidence = geography_match and service_scope_match and not synthetic and not blocked_geo and source_type in {
                "standard_charges_csv",
                "duke_standard_charges_csv",
                "shoppable_services_filters_json",
                "local_transparency_file",
                "public_url_manifest",
            }
            allowed_as_context = geography_match and not blocked_geo
            score = 1.0 if allowed_as_evidence else (0.4 if allowed_as_context else 0.0)
            reason = "evidence" if allowed_as_evidence else ("context_only" if allowed_as_context else "irrelevant")
            evaluations.append(
                {
                    "source_name": source_name,
                    "relevance_score": score,
                    "geography_match": geography_match,
                    "service_scope_match": service_scope_match,
                    "source_type": source_type,
                    "allowed_as_evidence": allowed_as_evidence,
                    "allowed_as_context": allowed_as_context,
                    "reason": reason if not blocked_geo else f"blocked_by:{blocked_geo}",
                    "synthetic": synthetic,
                }
            )
            if blocked_geo or not allowed_as_context:
                violations.append(
                    ValidationViolation(
                        artifact=source_name or "source_manifest",
                        matched_text=blocked_geo or source_geo or source_type,
                        reason=f"Source is not relevant evidence for {contract.get('geography', '')} {contract.get('service_focus', '')}".strip(),
                    )
                )
            if allowed_as_evidence:
                allowed_evidence_count += 1
        passed = allowed_evidence_count > 0 or any(bool(item.get("synthetic")) for item in evaluations)
        return ValidationResult(
            validator="SourceRelevanceValidator",
            passed=passed,
            severity="blocking",
            violations=violations[:80] if not passed or violations else [],
            score=1.0 if passed else 0.0,
            metadata={
                "source_evaluations": evaluations,
                "allowed_evidence_count": allowed_evidence_count,
                "synthetic_only": bool(evaluations) and allowed_evidence_count == 0 and all(bool(item.get("synthetic")) for item in evaluations),
            },
        )


class ArtifactContaminationValidator:
    def validate(
        self,
        *,
        contract: Dict[str, Any],
        artifacts: Dict[str, str],
        geography_result: ValidationResult,
        service_result: ValidationResult,
        source_result: ValidationResult,
    ) -> ValidationResult:
        violations: List[ValidationViolation] = []
        statuses: Dict[str, str] = {}
        invalid_artifacts: List[str] = []
        for key, raw_path in (artifacts or {}).items():
            path = Path(str(raw_path))
            text = _read_textish_artifact(path)
            title_probe = f"{path.name} {text[:1200]}"
            invalid = False
            if not str(raw_path).strip():
                violations.append(ValidationViolation(artifact=key, matched_text="", reason="Empty artifact path"))
                invalid = True
            if key not in {"task_contract_json", "geography_validation_report_md"} and not path.exists():
                violations.append(ValidationViolation(artifact=key, matched_text=path.name, reason="Artifact file is missing"))
                invalid = True
            if key.endswith("_md") or path.suffix.lower() in {".md", ".html"}:
                if not text.strip():
                    violations.append(ValidationViolation(artifact=key, matched_text=path.name, reason="Artifact is empty or placeholder-only"))
                    invalid = True
            if str(contract.get("geography", "")).strip() and path.suffix.lower() in {".md", ".html", ".xlsx"}:
                if not _geography_reflected_in_probe(title_probe, str(contract.get("geography", ""))):
                    violations.append(ValidationViolation(artifact=key, matched_text=path.name, reason="Artifact title/content does not reflect requested geography"))
                    invalid = True
            if geography_result.violations and any(v.artifact == key for v in geography_result.violations):
                invalid = True
            if service_result.violations and any(v.artifact == key for v in service_result.violations):
                invalid = True
            if not source_result.passed and key not in {"task_contract_json", "geography_validation_report_md"}:
                invalid = True
            statuses[key] = "quarantined" if invalid else "valid"
            if invalid:
                invalid_artifacts.append(key)
        return ValidationResult(
            validator="ArtifactContaminationValidator",
            passed=not invalid_artifacts,
            severity="blocking",
            violations=violations[:120],
            score=1.0 if not invalid_artifacts else 0.0,
            metadata={"artifact_statuses": statuses, "invalid_artifacts": invalid_artifacts},
        )


class FinalOutputGate:
    def evaluate(
        self,
        *,
        validation_results: Iterable[ValidationResult],
        completion_passed: bool,
        required_artifacts_exist: bool,
    ) -> FinalOutputGateResult:
        blocking_failures: List[str] = []
        required_repairs: List[str] = []
        issue_count = 0
        for result in validation_results:
            if not result.passed and result.severity == "blocking":
                blocking_failures.append(result.validator)
                issue_count += len(result.violations)
                if result.validator == "GeographyValidator":
                    required_repairs.append("Quarantine wrong-geography sources and rebuild the source manifest for the requested market.")
                elif result.validator == "ServiceScopeValidator":
                    required_repairs.append("Filter the dataset to the requested service scope and regenerate stakeholder artifacts.")
                elif result.validator == "SourceRelevanceValidator":
                    required_repairs.append("Remove irrelevant sources from evidence and rebuild with locally relevant or clearly labeled synthetic inputs.")
                elif result.validator == "ArtifactContaminationValidator":
                    required_repairs.append("Quarantine invalid artifacts and regenerate the final package from clean inputs.")
        if not completion_passed:
            blocking_failures.append("CompletionCritic")
            required_repairs.append("Generate the missing validated deliverables before completion.")
        if not required_artifacts_exist:
            blocking_failures.append("RequiredArtifacts")
            required_repairs.append("Restore required artifacts or return a blocked/partial result honestly.")
        return FinalOutputGateResult(
            passed=not blocking_failures,
            severity="blocking" if blocking_failures else "low",
            blocking_failures=blocking_failures,
            required_repairs=required_repairs,
            issue_count=issue_count,
        )
